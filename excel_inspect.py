import argparse
from os import remove
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description="Excel differeces finder",
                                 formatter_class=argparse.ArgumentDefaultsHelpFormatter)
parser.add_argument("-a", "--book-a", help="Excel A")
parser.add_argument("-b", "--book-b", default="", help="Excel B")
parser.add_argument("-v", "--verbose", action="store_true", help="Show results in console too")
parser.add_argument("-o", "--output-file", default="", help="Output filename with results")
parser.add_argument("-f", "--force-overwrite", action="store_true", help="Force overwrite output file if already exists")

args = parser.parse_args()
config = vars(args)

# setup vars
xls_file1 = config["book_a"]
xls_file2 = config["book_b"]
ofilename = config["output_file"]
verbose = config["verbose"]
overwrite = config["force_overwrite"]

# process

coll_A = []
coll_B = []
coll_A_B = []
coll_B_A = []
coll_sames = []

book1 = load_workbook(xls_file1)
book1.sheetnames
sheet1 = book1.active

if xls_file2:
  book2 = load_workbook(xls_file2)
  book2.sheetnames
  sheet2 = book2.active

  for col in sheet1['A']:
    if col.value:
      coll_A.append(col.value)

  for col in sheet2['A']:
    if col.value:
      coll_B.append(col.value)
else:

  for col in sheet1['A']:
    if col.value:
      coll_A.append(col.value)

  for col in sheet1['B']:
    if col.value:
      coll_B.append(col.value)

coll_A_B = list(set(coll_A).difference(set(coll_B)))
coll_B_A = list(set(coll_B).difference(set(coll_A)))
coll_sames = list(set(coll_A).intersection(coll_B))

if verbose:
  print(f"\nset A ({len(coll_A)} items)")
  print(coll_A)
  print(f"\nset B ({len(coll_B)} items)")
  print(coll_B)
  print("\nDifference A - B" )
  print(coll_A_B)
  print("\nDifference B - A")
  print(coll_B_A)
  print(f"\nEquality A = B ({len(coll_sames)} items)")
  print(coll_sames)

if ofilename != "":
  if ofilename.endswith('.txt'):
    if os.path.isfile(ofilename) and not overwrite:
      print(f"\n\nFile {ofilename} already exists. Use -f option or delete it first !!!")
      
    else:
      if os.path.isfile(ofilename):
          remove(ofilename)   

      file = open(ofilename, "w+")
      
      file.write(f"Difference A - B ({len(coll_A_B)} items):\n")
      file.write("----------------------------\n\n")
      for value in coll_A_B:
        file.write(f"{str(value)}\n")
      file.write("\n\n")

      file.write(f"Difference B - A ({len(coll_sames)} items):\n")
      file.write("----------------------------\n\n")
      for value in coll_B_A:
        file.write(f"{str(value)}\n")
      file.write("\n\n")

      file.write(f"Equality A = B ({len(coll_B_A)} items):\n")
      file.write("----------------------------\n\n")
      for value in coll_sames:
        file.write(f"{str(value)}\n")

      file.close()

  elif ofilename.endswith('.xls') or ofilename.endswith('.xlsx'):
      if os.path.isfile(ofilename) and not overwrite:
        print(f"\n\nFile {ofilename} already exists. Use -f option or delete it first !!!")
      else:
        if os.path.isfile(ofilename):
          remove(ofilename) 

        outbook = Workbook()
        
        outsheet = outbook.active
        outsheet.title = "Results"

        # Difference A - B
        nrow = 1
        outsheet.cell(row=nrow, column=1).value = "Difference A - B"
        for value in coll_A_B:
          nrow += 1
          outsheet.cell(row=nrow, column=1).value = str(value)

        # Difference B - A
        nrow = 1
        outsheet.cell(row=nrow, column=2).value = "Difference B - A"
        for value in coll_B_A:
          nrow += 1
          outsheet.cell(row=nrow, column=2).value = str(value)
        
        # # Equality A = B
        nrow = 1
        outsheet.cell(row=nrow, column=3).value = "Equality A = B"
        for value in coll_sames:
          nrow += 1
          outsheet.cell(row=nrow, column=3).value = str(value)
        
        outbook.save(filename=ofilename)
